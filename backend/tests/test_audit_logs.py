"""
P18 操作日志增强测试
- 按模块/操作类型/操作人/日期范围筛选
- 导出 Excel
- 权限校验（老师/管理员）
"""
import pytest
from openpyxl import load_workbook

from apps.audit.models import OperationLog


def extract_data(response):
    """从统一响应格式中提取 data"""
    data = response.json()
    if isinstance(data, dict) and 'code' in data:
        return data.get('data', data)
    return data


def get_results(data):
    """从分页或非分页数据中提取结果列表"""
    if isinstance(data, dict):
        return data.get('results', data)
    return data


@pytest.fixture
def make_log(db, make_user):
    """创建操作日志的工厂函数"""
    counter = [0]

    def _make(
        operator=None,
        operation_type=OperationLog.OperationType.CREATE,
        module='projects',
        object_type='Project',
        object_id='1',
        description='测试操作',
        request_method='POST',
        request_path='/api/v1/projects/',
        is_success=True,
        created_at=None,
        **extra,
    ):
        counter[0] += 1
        operator = operator or make_user(
            email=f'log_op{counter[0]}@test.com',
            name=f'操作员{counter[0]}',
            global_role='teacher',
        )
        log = OperationLog.objects.create(
            operator=operator,
            operation_type=operation_type,
            module=module,
            object_type=object_type,
            object_id=object_id,
            description=description,
            request_method=request_method,
            request_path=request_path,
            is_success=is_success,
            **extra,
        )
        if created_at is not None:
            OperationLog.objects.filter(id=log.id).update(created_at=created_at)
            log.refresh_from_db()
        return log

    return _make


@pytest.mark.api
@pytest.mark.django_db
class TestOperationLogFilter:
    """操作日志筛选测试"""

    def test_filter_by_module(self, teacher_client, make_log):
        """按模块筛选"""
        make_log(module='projects', object_id='101')
        make_log(module='finance', object_id='202')
        resp = teacher_client.get('/api/v1/audit/operation-logs/', {'module': 'projects'})
        assert resp.status_code == 200
        object_ids = [l['object_id'] for l in get_results(extract_data(resp))]
        assert '101' in object_ids
        assert '202' not in object_ids

    def test_filter_by_operation_type(self, teacher_client, make_log):
        """按操作类型筛选"""
        make_log(operation_type=OperationLog.OperationType.CREATE, object_id='301')
        make_log(operation_type=OperationLog.OperationType.DELETE, object_id='302')
        resp = teacher_client.get('/api/v1/audit/operation-logs/', {'operation_type': 'delete'})
        assert resp.status_code == 200
        object_ids = [l['object_id'] for l in get_results(extract_data(resp))]
        assert '302' in object_ids
        assert '301' not in object_ids

    def test_filter_by_operator(self, teacher_client, make_log, make_user):
        """按操作人筛选"""
        op1 = make_user(email='op1@test.com', name='操作员1', global_role='member')
        op2 = make_user(email='op2@test.com', name='操作员2', global_role='member')
        make_log(operator=op1, object_id='401')
        make_log(operator=op2, object_id='402')
        resp = teacher_client.get('/api/v1/audit/operation-logs/', {'operator': op1.id})
        assert resp.status_code == 200
        results = get_results(extract_data(resp))
        operators = [l['operator'] for l in results]
        assert op1.id in operators
        assert op2.id not in operators

    def test_filter_by_date_range(self, teacher_client, make_log):
        """按日期范围筛选"""
        from django.utils import timezone
        from datetime import timedelta
        now = timezone.now()
        old = now - timedelta(days=10)
        make_log(object_id='501', created_at=old)
        make_log(object_id='502', created_at=now)
        # 仅查最近 5 天
        start = (now - timedelta(days=5)).strftime('%Y-%m-%d')
        resp = teacher_client.get('/api/v1/audit/operation-logs/', {'start_date': start})
        assert resp.status_code == 200
        object_ids = [l['object_id'] for l in get_results(extract_data(resp))]
        assert '502' in object_ids
        assert '501' not in object_ids

    def test_combined_filters(self, teacher_client, make_log):
        """组合筛选"""
        make_log(module='projects', operation_type=OperationLog.OperationType.CREATE, object_id='601')
        make_log(module='projects', operation_type=OperationLog.OperationType.DELETE, object_id='602')
        make_log(module='finance', operation_type=OperationLog.OperationType.CREATE, object_id='603')
        resp = teacher_client.get('/api/v1/audit/operation-logs/', {
            'module': 'projects',
            'operation_type': 'create',
        })
        assert resp.status_code == 200
        object_ids = [l['object_id'] for l in get_results(extract_data(resp))]
        assert '601' in object_ids
        assert '602' not in object_ids
        assert '603' not in object_ids


@pytest.mark.api
@pytest.mark.django_db
class TestOperationLogExport:
    """操作日志导出 Excel 测试"""

    def test_export_returns_excel(self, teacher_client, make_log):
        """导出返回 Excel 文件"""
        make_log(description='导出测试1')
        make_log(description='导出测试2')
        resp = teacher_client.get('/api/v1/audit/operation-logs/export/')
        assert resp.status_code == 200
        assert 'spreadsheet' in resp['Content-Type']
        assert 'attachment' in resp['Content-Disposition']
        assert resp['Content-Disposition'].endswith('.xlsx"')

        # 解析 Excel 内容
        wb = load_workbook(io_bytes(resp.content))
        ws = wb.active
        # 表头 + 2 行数据
        assert ws.max_row == 3
        headers = [c.value for c in ws[1]]
        assert '操作人' in headers
        assert '操作类型' in headers
        assert '操作模块' in headers

    def test_export_respects_filters(self, teacher_client, make_log):
        """导出应用筛选条件"""
        make_log(module='projects', description='项目日志')
        make_log(module='finance', description='经费日志')
        resp = teacher_client.get('/api/v1/audit/operation-logs/export/', {'module': 'projects'})
        assert resp.status_code == 200
        wb = load_workbook(io_bytes(resp.content))
        ws = wb.active
        # 表头 + 1 行数据（仅项目日志）
        assert ws.max_row == 2
        # values_only=True 时每行为值元组，row[12] 即「操作描述」列
        descs = [row[12] for row in ws.iter_rows(min_row=2, values_only=True)]
        assert '项目日志' in descs
        assert '经费日志' not in descs

    def test_export_empty(self, teacher_client):
        """无数据时导出仅含表头"""
        resp = teacher_client.get('/api/v1/audit/operation-logs/export/')
        assert resp.status_code == 200
        wb = load_workbook(io_bytes(resp.content))
        ws = wb.active
        assert ws.max_row == 1  # 仅表头

    def test_export_member_forbidden(self, member_client, make_log):
        """普通成员不可导出"""
        make_log(description='日志')
        resp = member_client.get('/api/v1/audit/operation-logs/export/')
        assert resp.status_code == 403


@pytest.mark.api
@pytest.mark.django_db
class TestOperationLogPermission:
    """操作日志权限测试"""

    def test_member_cannot_list(self, member_client, make_log):
        """普通成员不可查看操作日志"""
        make_log(description='日志')
        resp = member_client.get('/api/v1/audit/operation-logs/')
        assert resp.status_code == 403

    def test_unauthenticated_cannot_list(self, api_client, make_log):
        """未登录不可查看"""
        make_log(description='日志')
        resp = api_client.get('/api/v1/audit/operation-logs/')
        assert resp.status_code == 401

    def test_admin_can_list(self, admin_client, make_log):
        """管理员可查看"""
        make_log(description='管理员可见')
        resp = admin_client.get('/api/v1/audit/operation-logs/')
        assert resp.status_code == 200


def io_bytes(content):
    """将 bytes 包装为 BytesIO 以便 openpyxl 读取"""
    import io
    return io.BytesIO(content)
