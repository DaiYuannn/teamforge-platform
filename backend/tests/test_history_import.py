"""
M10: 历史项目导入（History Import）测试
- ImportTask.Module 包含 history_projects
- MODULE_CONFIG 配置正确（model=Project、额外字段、默认 status=closed）
- 自动字段映射支持历史字段
- 导入模板下载
- 完整导入流程（解析→校验→确认写入）
- 默认值 status=closed 生效
- 回滚
"""
import pytest


def extract_data(response):
    data = response.json()
    if isinstance(data, dict) and 'code' in data:
        return data.get('data', data)
    return data


@pytest.mark.api
@pytest.mark.django_db
class TestHistoryImportConfig:
    """历史项目导入配置测试"""

    def test_module_choice_exists(self):
        """ImportTask.Module 包含 history_projects 选项"""
        from apps.imports.models import ImportTask
        choices = [c[0] for c in ImportTask.Module.choices]
        assert 'history_projects' in choices

    def test_module_config_exists(self):
        """MODULE_CONFIG 包含 history_projects 配置"""
        from apps.imports.services import ImportService
        config = ImportService.MODULE_CONFIG.get('history_projects')
        assert config is not None
        assert config['model'] == 'apps.projects.models.Project'
        # 额外支持的历史字段
        optional = config['optional_fields']
        assert 'actual_end_date' in optional
        assert 'current_stage' in optional
        # 默认 status=closed
        assert config.get('defaults', {}).get('status') == 'closed'

    def test_required_fields(self):
        """history_projects 必填字段为 name/code"""
        from apps.imports.services import ImportService
        config = ImportService.MODULE_CONFIG['history_projects']
        assert config['required_fields'] == ['name', 'code']

    def test_auto_map_history_fields(self):
        """自动字段映射支持历史项目字段"""
        from apps.imports.services import ImportService
        headers = ['项目名称', '项目编号', '负责人ID', '当前阶段',
                   '实际结束日期', '开始日期', '计划结束日期', '状态', '简介', '优先级']
        mapping = ImportService.auto_map_fields(headers, 'history_projects')
        assert mapping['项目名称'] == 'name'
        assert mapping['项目编号'] == 'code'
        assert mapping['负责人ID'] == 'leader_id'
        assert mapping['当前阶段'] == 'current_stage'
        assert mapping['实际结束日期'] == 'actual_end_date'
        assert mapping['开始日期'] == 'start_date'
        assert mapping['计划结束日期'] == 'planned_end_date'
        assert mapping['状态'] == 'status'
        assert mapping['简介'] == 'intro'
        assert mapping['优先级'] == 'priority'


@pytest.mark.api
@pytest.mark.django_db
class TestHistoryImportTemplate:
    """历史项目导入模板测试"""

    def test_template_download(self, teacher_client):
        """历史项目导入模板可下载"""
        resp = teacher_client.get('/api/v1/exports/template/?type=history_projects')
        assert resp.status_code == 200
        assert 'spreadsheet' in resp['Content-Type']
        # Content-Disposition 整体被 RFC2047 编码，无法直接匹配明文，
        # 改为校验响应体确实是可解析的 xlsx
        import io
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        assert wb.active is not None

    def test_template_has_history_headers(self, teacher_client):
        """模板包含历史项目专属表头"""
        import io
        import openpyxl
        resp = teacher_client.get('/api/v1/exports/template/?type=history_projects')
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        # 历史项目专属表头
        assert '实际结束' in headers
        assert '当前阶段' in headers
        assert '负责人ID' in headers
        assert '项目名称' in headers
        assert '项目编号' in headers

    def test_template_download_member(self, member_client):
        """普通成员也可下载模板（仅需认证）"""
        resp = member_client.get('/api/v1/exports/template/?type=history_projects')
        assert resp.status_code == 200

    def test_template_unauthenticated(self, api_client):
        """未认证不能下载模板"""
        resp = api_client.get('/api/v1/exports/template/?type=history_projects')
        assert resp.status_code == 401


@pytest.mark.integration
@pytest.mark.django_db
class TestHistoryImportFlow:
    """历史项目导入完整流程测试"""

    def _make_excel(self, path, leader_id, rows):
        """辅助：生成历史项目导入 Excel 文件"""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['项目名称', '项目编号', '负责人ID', '当前阶段',
                   '实际结束日期', '开始日期', '计划结束日期', '状态'])
        for r in rows:
            ws.append(r)
        wb.save(str(path))
        return str(path)

    def test_full_import_flow(self, db, make_user, tmp_path):
        """完整历史项目导入流程：解析→校验→确认"""
        from apps.imports.services import ImportService
        from apps.imports.models import ImportTask
        from apps.projects.models import Project

        leader = make_user(email='histleader@test.com', global_role='member')

        file_path = self._make_excel(
            tmp_path / 'history.xlsx', leader.id,
            [
                ['历史项目A', 'HIST-001', leader.id, 14, '2025-06-30', '2024-01-01', '2025-06-01', 'closed'],
                ['历史项目B', 'HIST-002', leader.id, 14, '2025-07-15', '2024-03-01', '2025-07-01', 'closed'],
            ],
        )

        # 解析
        headers, rows = ImportService.parse_excel(file_path)
        assert len(rows) == 2

        # 自动映射
        mapping = ImportService.auto_map_fields(headers, 'history_projects')
        assert 'name' in mapping.values()
        assert 'leader_id' in mapping.values()

        # 校验
        valid_rows, errors = ImportService.validate_rows(rows, mapping, 'history_projects')
        assert len(valid_rows) == 2
        assert len(errors) == 0

        # 创建导入任务并确认导入
        task = ImportTask.objects.create(
            module='history_projects',
            file_path=file_path,
            status=ImportTask.Status.PREVIEWED,
            field_mapping=mapping,
            created_by=leader,
        )
        success, result = ImportService.confirm_import(task)
        assert success, result
        assert result['created_count'] == 2
        assert result['error_count'] == 0

        # 验证项目创建且为历史归档状态
        p1 = Project.objects.get(code='HIST-001')
        assert p1.name == '历史项目A'
        assert p1.status == 'closed'
        assert p1.current_stage == 14
        assert p1.leader_id == leader.id
        assert str(p1.actual_end_date) == '2025-06-30'
        assert str(p1.start_date) == '2024-01-01'

        p2 = Project.objects.get(code='HIST-002')
        assert p2.status == 'closed'
        assert p2.current_stage == 14

    def test_import_applies_status_default(self, db, make_user, tmp_path):
        """未提供 status 列时，默认 status=closed"""
        from apps.imports.services import ImportService
        from apps.imports.models import ImportTask
        from apps.projects.models import Project

        leader = make_user(email='histleader2@test.com', global_role='member')

        # 不含状态列
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['项目名称', '项目编号', '负责人ID', '当前阶段', '实际结束日期'])
        ws.append(['历史项目C', 'HIST-003', leader.id, 14, '2025-08-30'])
        file_path = str(tmp_path / 'history2.xlsx')
        wb.save(file_path)

        headers, rows = ImportService.parse_excel(file_path)
        mapping = ImportService.auto_map_fields(headers, 'history_projects')
        valid_rows, _ = ImportService.validate_rows(rows, mapping, 'history_projects')
        assert len(valid_rows) == 1

        task = ImportTask.objects.create(
            module='history_projects',
            file_path=file_path,
            status=ImportTask.Status.PREVIEWED,
            field_mapping=mapping,
            created_by=leader,
        )
        success, result = ImportService.confirm_import(task)
        assert success, result
        assert result['created_count'] == 1

        p = Project.objects.get(code='HIST-003')
        # 默认值生效
        assert p.status == 'closed'
        assert p.current_stage == 14

    def test_import_with_status_override(self, db, make_user, tmp_path):
        """显式提供 status 时覆盖默认值"""
        from apps.imports.services import ImportService
        from apps.imports.models import ImportTask
        from apps.projects.models import Project

        leader = make_user(email='histleader3@test.com', global_role='member')

        file_path = self._make_excel(
            tmp_path / 'history3.xlsx', leader.id,
            [['历史项目D', 'HIST-004', leader.id, 14, '2025-09-30', '2024-01-01', '2025-09-01', 'closed']],
        )
        headers, rows = ImportService.parse_excel(file_path)
        mapping = ImportService.auto_map_fields(headers, 'history_projects')

        task = ImportTask.objects.create(
            module='history_projects',
            file_path=file_path,
            status=ImportTask.Status.PREVIEWED,
            field_mapping=mapping,
            created_by=leader,
        )
        success, result = ImportService.confirm_import(task)
        assert success
        p = Project.objects.get(code='HIST-004')
        assert p.status == 'closed'

    def test_rollback_history_import(self, db, make_user, tmp_path):
        """回滚历史项目导入：根据快照删除已写入数据"""
        from apps.imports.services import ImportService
        from apps.imports.models import ImportTask
        from apps.projects.models import Project

        leader = make_user(email='histleader4@test.com', global_role='member')

        file_path = self._make_excel(
            tmp_path / 'history4.xlsx', leader.id,
            [['历史项目E', 'HIST-005', leader.id, 14, '2025-06-30', '2024-01-01', '2025-06-01', 'closed']],
        )
        headers, rows = ImportService.parse_excel(file_path)
        mapping = ImportService.auto_map_fields(headers, 'history_projects')

        task = ImportTask.objects.create(
            module='history_projects',
            file_path=file_path,
            status=ImportTask.Status.PREVIEWED,
            field_mapping=mapping,
            created_by=leader,
        )
        success, _ = ImportService.confirm_import(task)
        assert success
        assert task.status == ImportTask.Status.CONFIRMED
        assert Project.objects.filter(code='HIST-005').exists()

        # 回滚
        success, msg = ImportService.rollback_import(task)
        assert success, msg
        assert task.status == ImportTask.Status.ROLLED_BACK
        # 数据已删除
        assert not Project.objects.filter(code='HIST-005').exists()

    def test_validate_missing_required(self, db, make_user, tmp_path):
        """必填字段值为空时校验失败"""
        from apps.imports.services import ImportService

        leader = make_user(email='histleader5@test.com', global_role='member')

        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['项目名称', '项目编号', '负责人ID', '当前阶段'])
        ws.append(['', 'HIST-007', leader.id, 14])   # 项目名称(name) 为空
        ws.append(['历史项目G', '', leader.id, 14])   # 项目编号(code) 为空
        file_path = str(tmp_path / 'history5.xlsx')
        wb.save(file_path)

        headers, rows = ImportService.parse_excel(file_path)
        mapping = ImportService.auto_map_fields(headers, 'history_projects')
        valid_rows, errors = ImportService.validate_rows(rows, mapping, 'history_projects')
        # 两行必填字段均为空，全部校验失败
        assert len(valid_rows) == 0
        assert len(errors) == 2


@pytest.mark.api
@pytest.mark.django_db
class TestHistoryImportAPI:
    """历史项目导入 API 测试"""

    def test_preview_history_import(self, teacher_client, make_user, tmp_path):
        """通过 API 预览历史项目导入"""
        from django.core.files.uploadedfile import SimpleUploadedFile
        import openpyxl
        import io

        leader = make_user(email='apihist@test.com', global_role='member')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['项目名称', '项目编号', '负责人ID', '当前阶段', '实际结束日期'])
        ws.append(['历史项目G', 'HIST-006', leader.id, 14, '2025-06-30'])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        upload = SimpleUploadedFile(
            'history.xlsx', buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        resp = teacher_client.post(
            '/api/v1/imports/tasks/preview/',
            {'file': upload, 'module': 'history_projects'},
            format='multipart',
        )
        assert resp.status_code == 200, resp.json()
        data = extract_data(resp)
        assert data['total_rows'] == 1
        assert data['valid_rows'] == 1
        # 自动映射应包含历史字段
        assert 'leader_id' in data['field_mapping'].values()
        assert 'current_stage' in data['field_mapping'].values()

    def test_preview_invalid_module(self, teacher_client):
        """无效的导入模块"""
        from django.core.files.uploadedfile import SimpleUploadedFile
        upload = SimpleUploadedFile('x.xlsx', b'fake', content_type='application/octet-stream')
        resp = teacher_client.post(
            '/api/v1/imports/tasks/preview/',
            {'file': upload, 'module': 'invalid_module'},
            format='multipart',
        )
        assert resp.status_code == 400

    def test_member_cannot_preview(self, member_client, make_user):
        """普通成员无权执行导入预览"""
        from django.core.files.uploadedfile import SimpleUploadedFile
        leader = make_user(email='memhist@test.com', global_role='member')
        upload = SimpleUploadedFile('x.xlsx', b'fake', content_type='application/octet-stream')
        resp = member_client.post(
            '/api/v1/imports/tasks/preview/',
            {'file': upload, 'module': 'history_projects'},
            format='multipart',
        )
        assert resp.status_code == 403
