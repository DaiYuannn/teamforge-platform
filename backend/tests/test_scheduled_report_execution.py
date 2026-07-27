"""定时报表文件生成、执行审计、下载和调度领取测试。"""
from datetime import time, timedelta

import pytest
from django.utils import timezone
from openpyxl import load_workbook
from rest_framework.test import APIClient

from apps.exports.custom_report_models import CustomReport
from apps.exports.scheduled_report_models import (
    ScheduledReport,
    ScheduledReportExecution,
)
from apps.exports.scheduled_report_service import (
    claim_due_schedule_execution_ids,
    claim_due_schedule_ids,
    compute_next_run,
    execute_scheduled_report,
)
from apps.notifications.models import Notification
from apps.users.models import User, UserPreference


@pytest.fixture
def report_schedule(member_client, settings, tmp_path):
    settings.MEDIA_ROOT = str(tmp_path / 'media')
    report = CustomReport.objects.create(
        name='项目执行概览',
        description='用于测试实际文件生成',
        report_type='project',
        config={
            'data_source': 'project',
            'group_by': 'status',
            'chart_type': 'table',
        },
        created_by=member_client.user,
    )
    schedule = ScheduledReport.objects.create(
        report=report,
        created_by=member_client.user,
        frequency='daily',
        execution_time=time(9, 0),
        file_format='xlsx',
    )
    schedule.next_run = compute_next_run(schedule)
    schedule.save(update_fields=['next_run'])
    return schedule


@pytest.mark.django_db
class TestScheduledReportExecution:
    def test_generates_real_xlsx_and_execution_record(
        self,
        report_schedule,
        member_client,
        make_project,
    ):
        make_project(status='active')
        execution = execute_scheduled_report(
            report_schedule,
            user=member_client.user,
        )
        assert execution.status == ScheduledReport.RunStatus.SUCCESS
        assert execution.file_name.endswith('.xlsx')
        assert execution.file_size > 0
        assert execution.file.storage.exists(execution.file.name)
        with execution.file.open('rb') as source:
            workbook = load_workbook(source, read_only=True)
        assert {'概览', '明细'} <= set(workbook.sheetnames)
        assert Notification.objects.filter(
            recipient=member_client.user,
            related_object_type='scheduled_report',
        ).exists()

    def test_email_not_configured_is_partial_but_file_remains(
        self,
        report_schedule,
        member_client,
        make_user,
        settings,
    ):
        settings.EMAIL_HOST_USER = ''
        report_schedule.recipients.add(make_user(email='receiver@example.com'))
        execution = execute_scheduled_report(report_schedule, user=member_client.user)
        assert execution.status == ScheduledReport.RunStatus.PARTIAL
        assert (
            execution.delivery_status
            == ScheduledReportExecution.DeliveryStatus.NOT_CONFIGURED
        )
        assert execution.file

    def test_report_preferences_disable_inapp_and_email_delivery(
        self,
        report_schedule,
        member_client,
        settings,
    ):
        settings.EMAIL_HOST_USER = 'mailer@example.com'
        report_schedule.recipients.add(member_client.user)
        UserPreference.objects.create(
            user=member_client.user,
            notification_preferences={
                'categories': {'report': False},
                'channels': {'in_app': True, 'email': True},
            },
        )
        with pytest.MonkeyPatch.context() as monkeypatch:
            def fail_email_delivery(*args, **kwargs):
                pytest.fail('偏好关闭后不应创建邮件')

            monkeypatch.setattr(
                'apps.exports.scheduled_report_service.EmailMessage',
                fail_email_delivery,
            )
            execution = execute_scheduled_report(
                report_schedule,
                user=member_client.user,
            )

        assert execution.status == ScheduledReport.RunStatus.SUCCESS
        assert (
            execution.delivery_status
            == ScheduledReportExecution.DeliveryStatus.NOT_REQUESTED
        )
        assert not Notification.objects.filter(
            recipient=member_client.user,
            related_object_type='scheduled_report',
        ).exists()

    def test_due_schedule_is_claimed_once(self, report_schedule):
        report_schedule.next_run = timezone.now() - timedelta(minutes=1)
        report_schedule.save(update_fields=['next_run'])
        due_at = report_schedule.next_run
        assert claim_due_schedule_ids() == [report_schedule.id]
        assert claim_due_schedule_ids() == []
        report_schedule.refresh_from_db()
        assert report_schedule.next_run == due_at
        assert report_schedule.executions.filter(
            trigger=ScheduledReportExecution.Trigger.SCHEDULED,
            status=ScheduledReport.RunStatus.RUNNING,
        ).count() == 1

    def test_claimed_execution_advances_schedule_only_after_success(
        self, report_schedule
    ):
        due_at = timezone.now() - timedelta(minutes=1)
        report_schedule.next_run = due_at
        report_schedule.save(update_fields=['next_run'])

        execution_id = claim_due_schedule_execution_ids()[0]
        report_schedule.refresh_from_db()
        assert report_schedule.next_run == due_at

        claimed = ScheduledReportExecution.objects.get(pk=execution_id)
        execution = execute_scheduled_report(
            report_schedule,
            trigger=ScheduledReportExecution.Trigger.SCHEDULED,
            execution=claimed,
        )

        report_schedule.refresh_from_db()
        assert execution.pk == execution_id
        assert execution.status == ScheduledReport.RunStatus.SUCCESS
        assert report_schedule.next_run > timezone.now()

    def test_failed_scheduled_execution_is_persisted_and_retried_soon(
        self, report_schedule, monkeypatch
    ):
        report_schedule.next_run = timezone.now() - timedelta(minutes=1)
        report_schedule.save(update_fields=['next_run'])
        execution_id = claim_due_schedule_execution_ids()[0]
        claimed = ScheduledReportExecution.objects.get(pk=execution_id)
        before = timezone.now()

        def fail_generation(schedule):
            raise RuntimeError('temporary generation failure')

        monkeypatch.setattr(
            'apps.exports.scheduled_report_service.generate_report_file',
            fail_generation,
        )
        execution = execute_scheduled_report(
            report_schedule,
            trigger=ScheduledReportExecution.Trigger.SCHEDULED,
            execution=claimed,
        )

        report_schedule.refresh_from_db()
        assert execution.status == ScheduledReport.RunStatus.FAILED
        assert execution.finished_at is not None
        assert 'temporary generation failure' in execution.error
        assert before + timedelta(minutes=4, seconds=55) < report_schedule.next_run
        assert report_schedule.next_run < before + timedelta(minutes=5, seconds=5)

    def test_stale_running_execution_is_failed_and_reclaimed(
        self, report_schedule, settings
    ):
        settings.SCHEDULED_REPORT_EXECUTION_TIMEOUT_SECONDS = 60
        report_schedule.next_run = timezone.now() - timedelta(minutes=5)
        report_schedule.save(update_fields=['next_run'])
        stale = ScheduledReportExecution.objects.create(
            schedule=report_schedule,
            trigger=ScheduledReportExecution.Trigger.SCHEDULED,
        )
        ScheduledReportExecution.objects.filter(pk=stale.pk).update(
            started_at=timezone.now() - timedelta(minutes=2)
        )

        claimed_ids = claim_due_schedule_execution_ids()

        stale.refresh_from_db()
        assert stale.status == ScheduledReport.RunStatus.FAILED
        assert stale.finished_at is not None
        assert len(claimed_ids) == 1
        assert claimed_ids[0] != stale.pk

    def test_worker_error_does_not_stop_other_due_reports(
        self, report_schedule, monkeypatch
    ):
        from apps.exports import tasks as report_tasks

        now = timezone.now()
        report_schedule.next_run = now - timedelta(minutes=2)
        report_schedule.save(update_fields=['next_run'])
        second = ScheduledReport.objects.create(
            report=report_schedule.report,
            created_by=report_schedule.created_by,
            frequency='daily',
            execution_time=time(10, 0),
            file_format='xlsx',
            next_run=now - timedelta(minutes=1),
        )
        real_execute = execute_scheduled_report
        calls = []

        def fail_first(schedule, **kwargs):
            calls.append(schedule.pk)
            if schedule.pk == report_schedule.pk:
                raise RuntimeError('worker interruption')
            return real_execute(schedule, **kwargs)

        monkeypatch.setattr(report_tasks, 'execute_scheduled_report', fail_first)
        result = report_tasks.run_due_scheduled_reports()

        assert calls == [report_schedule.pk, second.pk]
        assert result['executed'] == 2
        assert ScheduledReportExecution.objects.filter(
            schedule=report_schedule,
            status=ScheduledReport.RunStatus.FAILED,
        ).exists()
        failed = ScheduledReportExecution.objects.get(
            schedule=report_schedule,
            status=ScheduledReport.RunStatus.FAILED,
        )
        report_schedule.refresh_from_db()
        assert failed.finished_at is not None
        assert 'worker interruption' in failed.error
        assert report_schedule.last_status == ScheduledReport.RunStatus.FAILED
        assert 'worker interruption' in report_schedule.last_error
        assert now + timedelta(minutes=4, seconds=55) < report_schedule.next_run
        assert report_schedule.next_run < now + timedelta(minutes=5, seconds=5)
        assert ScheduledReportExecution.objects.filter(
            schedule=second,
            status=ScheduledReport.RunStatus.SUCCESS,
        ).exists()

    def test_weekly_rule_uses_selected_weekday(self, report_schedule):
        report_schedule.frequency = ScheduledReport.Frequency.WEEKLY
        report_schedule.weekday = 4
        report_schedule.execution_time = time(15, 30)
        next_run = compute_next_run(report_schedule)
        local = timezone.localtime(next_run)
        assert local.weekday() == 4
        assert local.hour == 15
        assert local.minute == 30

    def test_api_run_and_download(
        self,
        report_schedule,
        member_client,
        make_project,
    ):
        make_project(status='closed')
        response = member_client.post(
            f'/api/v1/exports/scheduled-reports/{report_schedule.id}/run_now/'
        )
        assert response.status_code == 200, response.json()
        data = response.json()['data']
        assert data['status'] == 'success'
        assert data['download_url']
        download_response = member_client.get(
            f'/api/v1/exports/scheduled-reports/{report_schedule.id}/'
            f'executions/{data["id"]}/download/'
        )
        assert download_response.status_code == 200
        assert 'attachment' in download_response['Content-Disposition']

    def test_member_cannot_schedule_another_members_report(
        self, member_client, make_user
    ):
        owner = make_user(email='scheduled-report-owner@test.com')
        report = CustomReport.objects.create(
            name='他人报表',
            report_type='summary',
            config={'data_source': 'project'},
            created_by=owner,
        )

        response = member_client.post(
            '/api/v1/exports/scheduled-reports/',
            {
                'report': report.id,
                'frequency': 'daily',
                'execution_time': '09:00:00',
                'file_format': 'xlsx',
            },
            format='json',
        )

        assert response.status_code == 400, response.json()
        assert not ScheduledReport.objects.filter(
            report=report,
            created_by=member_client.user,
        ).exists()

    def test_external_or_exited_recipient_is_rejected(
        self, member_client, make_user
    ):
        report = CustomReport.objects.create(
            name='接收人边界测试',
            report_type='summary',
            config={'data_source': 'project'},
            created_by=member_client.user,
        )
        external = make_user(
            email='scheduled-external-recipient@test.com',
            membership_status=User.MembershipStatus.EXTERNAL,
        )

        response = member_client.post(
            '/api/v1/exports/scheduled-reports/',
            {
                'report': report.id,
                'recipient_ids': [external.id],
                'frequency': 'daily',
                'execution_time': '09:00:00',
                'file_format': 'xlsx',
            },
            format='json',
        )

        assert response.status_code == 400, response.json()

    def test_recipient_can_download_but_cannot_manage_schedule(
        self, report_schedule, member_client, make_user, make_project
    ):
        recipient = make_user(email='scheduled-recipient@test.com')
        report_schedule.recipients.add(recipient)
        make_project()
        execution = execute_scheduled_report(
            report_schedule,
            user=member_client.user,
        )
        client = APIClient()
        client.force_authenticate(user=recipient)

        detail = client.get(
            f'/api/v1/exports/scheduled-reports/{report_schedule.id}/'
        )
        download = client.get(
            f'/api/v1/exports/scheduled-reports/{report_schedule.id}/'
            f'executions/{execution.id}/download/'
        )
        run_now = client.post(
            f'/api/v1/exports/scheduled-reports/{report_schedule.id}/run_now/'
        )
        activate = client.post(
            f'/api/v1/exports/scheduled-reports/{report_schedule.id}/activate/'
        )
        deactivate = client.post(
            f'/api/v1/exports/scheduled-reports/{report_schedule.id}/deactivate/'
        )
        updated = client.patch(
            f'/api/v1/exports/scheduled-reports/{report_schedule.id}/',
            {'file_format': 'pdf'},
            format='json',
        )
        deleted = client.delete(
            f'/api/v1/exports/scheduled-reports/{report_schedule.id}/'
        )

        assert detail.status_code == 200
        assert download.status_code == 200
        assert run_now.status_code == 403
        assert activate.status_code == 403
        assert deactivate.status_code == 403
        assert updated.status_code == 403
        assert deleted.status_code == 403

    def test_invalid_legacy_schedule_is_disabled_before_data_generation(
        self, member_client, make_user, settings, tmp_path
    ):
        settings.MEDIA_ROOT = str(tmp_path / 'media')
        external = make_user(
            email='scheduled-invalid-owner@test.com',
            membership_status=User.MembershipStatus.EXTERNAL,
        )
        report = CustomReport.objects.create(
            name='越权旧计划',
            report_type='summary',
            config={'data_source': 'finance'},
            created_by=member_client.user,
        )
        schedule = ScheduledReport.objects.create(
            report=report,
            created_by=external,
            frequency='daily',
            execution_time=time(9, 0),
            file_format='xlsx',
            is_active=True,
        )

        execution = execute_scheduled_report(schedule, user=external)

        schedule.refresh_from_db()
        assert execution.status == ScheduledReport.RunStatus.FAILED
        assert not execution.file
        assert '创建人不是有效内部成员' in execution.error
        assert schedule.is_active is False
