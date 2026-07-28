"""
Excel 导出服务
使用 openpyxl 生成各业务模块的 Excel 报表
所有导出接口直接返回文件流 HttpResponse
"""
import csv
import io
from decimal import Decimal

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.http import HttpResponse
from django.db.models import Q


# ============ 样式常量 ============
_HEADER_FONT = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
_HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
_CELL_FONT = Font(name='微软雅黑', size=10)
_HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)
_CELL_ALIGNMENT = Alignment(horizontal='left', vertical='center', wrap_text=True)
_THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)


def _apply_header_style(ws, headers, row=1):
    """写入并样式化标题行"""
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT
        cell.border = _THIN_BORDER
    ws.row_dimensions[row].height = 24


def _apply_body_style(ws, data_rows, start_row=2):
    """写入并样式化数据行"""
    for row_offset, row_data in enumerate(data_rows, start=start_row):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_offset, column=col_idx, value=value)
            cell.font = _CELL_FONT
            cell.alignment = _CELL_ALIGNMENT
            cell.border = _THIN_BORDER


def _auto_column_width(ws, headers, data_rows):
    """自适应列宽"""
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for row in data_rows:
            val = row[col_idx - 1] if col_idx - 1 < len(row) else ''
            max_len = max(max_len, len(str(val)) if val is not None else 0)
        # 中文宽度近似处理
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len * 2 + 4, 60)


def _wb_to_response(wb, filename):
    """
    Workbook 转 HttpResponse
    :param wb: openpyxl.Workbook
    :param filename: 下载文件名（不含扩展名）
    """
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    # 文件名兼容中文
    response['Content-Disposition'] = f"attachment; filename*=UTF-8''{filename}.xlsx"
    return response


def _task_export_queryset(project_id=None, filters=None, user=None):
    """复用任务列表的核心筛选语义生成导出数据集。"""
    from apps.tasks.models import Task

    filters = filters or {}
    tasks = (
        Task.objects.select_related(
            'project',
            'assignee',
            'creator',
            'reviewer',
        )
        .prefetch_related('collaborators')
    )
    if user is not None:
        from common.project_access import scope_project_queryset

        tasks = scope_project_queryset(
            tasks,
            user,
            project_lookup='project',
        )
    if project_id:
        tasks = tasks.filter(project_id=project_id)
    if filters.get('status'):
        tasks = tasks.filter(status=filters['status'])
    if filters.get('priority'):
        tasks = tasks.filter(priority=filters['priority'])
    if filters.get('assignee'):
        tasks = tasks.filter(assignee_id=filters['assignee'])
    search = str(filters.get('search') or '').strip()
    if search:
        tasks = tasks.filter(
            Q(title__icontains=search)
            | Q(description__icontains=search)
            | Q(project__name__icontains=search)
        )
    if filters.get('scope') == 'mine' and user:
        tasks = tasks.filter(
            Q(assignee=user)
            | Q(creator=user)
            | Q(reviewer=user)
            | Q(collaborators=user)
        )
    return tasks.distinct().order_by('-created_at')


_COMPETITION_EXPORT_HEADERS = [
    '所属项目', '项目编号', '比赛名称', '比赛类型', '级别', '主办单位',
    '状态', '当前阶段', '是否晋级', '是否获奖', '获奖等级',
    '报名日期', '材料提交截止', '网评日期', '答辩日期',
    '校赛日期', '市赛日期', '省赛日期', '国赛日期', '结果公布日期',
    '未晋级原因', '评审/答辩复盘', '改进建议', '创建时间',
]


def _format_date(value):
    return value.strftime('%Y-%m-%d') if value else ''


def _filtered_competitions(
    search='',
    level='',
    status='',
    project_id=None,
    user=None,
):
    """构建与比赛列表筛选语义一致的导出查询集。"""
    from apps.competitions.models import Competition

    queryset = Competition.objects.select_related('project').all().order_by('-created_at')
    if user is not None:
        from common.project_access import scope_project_queryset

        queryset = scope_project_queryset(
            queryset,
            user,
            project_lookup='project',
        )
    search = (search or '').strip()
    if search:
        queryset = queryset.filter(
            Q(name__icontains=search)
            | Q(organizer__icontains=search)
            | Q(project__name__icontains=search)
        )
    if level:
        queryset = queryset.filter(level=level)
    if status:
        queryset = queryset.filter(status=status)
    if project_id:
        project_id_value = str(project_id)
        queryset = (
            queryset.filter(project_id=int(project_id_value))
            if project_id_value.isdigit() and int(project_id_value) > 0
            else queryset.none()
        )
    return queryset


def _competition_export_rows(
    search='',
    level='',
    status='',
    project_id=None,
    user=None,
):
    rows = []
    for competition in _filtered_competitions(
        search,
        level,
        status,
        project_id,
        user,
    ):
        rows.append([
            competition.project.name if competition.project else '',
            competition.project.code if competition.project else '',
            competition.name,
            competition.comp_type,
            competition.get_level_display(),
            competition.organizer,
            competition.get_status_display(),
            competition.current_stage,
            '是' if competition.is_promoted else '否',
            '是' if competition.is_awarded else '否',
            competition.award_level,
            _format_date(competition.register_date),
            _format_date(competition.material_deadline),
            _format_date(competition.review_date),
            _format_date(competition.defense_date),
            _format_date(competition.school_date),
            _format_date(competition.city_date),
            _format_date(competition.province_date),
            _format_date(competition.national_date),
            _format_date(competition.result_date),
            competition.not_promoted_reason,
            competition.review_summary,
            competition.improvement_suggestion,
            competition.created_at.strftime('%Y-%m-%d %H:%M')
            if competition.created_at else '',
        ])
    return rows


class ExcelExportService:
    """Excel 导出服务"""

    @staticmethod
    def export_projects(user=None):
        """导出项目列表 Excel"""
        from apps.projects.models import Project
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '项目列表'
        headers = ['项目名称', '项目编号', '负责人', '当前阶段', '状态', '开始时间', '预计结束', '创建时间']
        _apply_header_style(ws, headers)

        projects = Project.objects.select_related('leader').all().order_by('-created_at')
        if user is not None:
            from common.project_access import scope_project_queryset

            projects = scope_project_queryset(projects, user, project_lookup='')
        data_rows = []
        for p in projects:
            data_rows.append([
                p.name,
                p.code,
                p.leader.name if p.leader else '',
                p.get_current_stage_display(),
                p.get_status_display(),
                p.start_date.strftime('%Y-%m-%d') if p.start_date else '',
                p.planned_end_date.strftime('%Y-%m-%d') if p.planned_end_date else '',
                p.created_at.strftime('%Y-%m-%d %H:%M') if p.created_at else '',
            ])
        _apply_body_style(ws, data_rows)
        _auto_column_width(ws, headers, data_rows)
        return _wb_to_response(wb, '项目列表')

    @staticmethod
    def export_finance_budget(user=None):
        """导出经费总表 Excel"""
        from apps.finance.models import FinanceBudget
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '经费总表'
        headers = ['项目名称', '项目编号', '奖金总额', '其他收入', '已用金额',
                   '待报销', '剩余金额', '经费状态', '统计周期', '更新时间']
        _apply_header_style(ws, headers)

        budgets = FinanceBudget.objects.select_related('project').all().order_by('-updated_at')
        if user is not None:
            from common.project_access import scope_project_queryset

            budgets = scope_project_queryset(
                budgets,
                user,
                project_lookup='project',
            )
        data_rows = []
        for b in budgets:
            data_rows.append([
                b.project.name if b.project else '',
                b.project.code if b.project else '',
                float(b.bonus_amount),
                float(b.other_income),
                float(b.used_amount),
                float(b.pending_reimbursement),
                float(b.remaining_amount),
                b.get_status_display(),
                b.period,
                b.updated_at.strftime('%Y-%m-%d %H:%M') if b.updated_at else '',
            ])
        _apply_body_style(ws, data_rows)
        _auto_column_width(ws, headers, data_rows)
        return _wb_to_response(wb, '经费总表')

    @staticmethod
    def export_finance_detail(project_id, user=None):
        """导出单项目经费明细 Excel"""
        from apps.finance.models import FinanceExpense
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '经费明细'
        headers = ['支出标题', '项目名称', '金额', '经办人', '支出日期',
                   '经费类别', '用途说明', '审核人', '创建时间']
        _apply_header_style(ws, headers)

        expenses = FinanceExpense.objects.select_related(
            'project', 'spender', 'reviewer'
        ).filter(project_id=project_id).order_by('-expense_date')
        if user is not None:
            from common.project_access import scope_project_queryset

            expenses = scope_project_queryset(
                expenses,
                user,
                project_lookup='project',
            )
        data_rows = []
        for e in expenses:
            data_rows.append([
                e.title,
                e.project.name if e.project else '',
                float(e.amount),
                e.spender.name if e.spender else '',
                e.expense_date.strftime('%Y-%m-%d') if e.expense_date else '',
                e.get_category_display(),
                e.purpose,
                e.reviewer.name if e.reviewer else '',
                e.created_at.strftime('%Y-%m-%d %H:%M') if e.created_at else '',
            ])
        _apply_body_style(ws, data_rows)
        _auto_column_width(ws, headers, data_rows)
        return _wb_to_response(wb, '经费明细')

    @staticmethod
    def export_tasks(project_id=None, filters=None, user=None):
        """按任务列表当前筛选导出任务清单 Excel。"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '任务清单'
        headers = [
            '任务标题', '所属项目', '指派给', '协作者', '审核人', '创建者',
            '状态', '优先级', '截止时间', '完成时间', '是否逾期',
            '延期原因', '完成说明', '创建时间',
        ]
        _apply_header_style(ws, headers)

        tasks = _task_export_queryset(project_id, filters, user)
        data_rows = []
        for t in tasks:
            data_rows.append([
                t.title,
                t.project.name if t.project else '',
                t.assignee.name if t.assignee else '',
                '、'.join(user.name for user in t.collaborators.all()),
                t.reviewer.name if t.reviewer else '',
                t.creator.name if t.creator else '',
                t.get_status_display(),
                t.get_priority_display(),
                t.deadline.strftime('%Y-%m-%d %H:%M') if t.deadline else '',
                t.completed_at.strftime('%Y-%m-%d %H:%M') if t.completed_at else '',
                '是' if t.is_overdue else '否',
                t.delay_reason,
                t.completion_note,
                t.created_at.strftime('%Y-%m-%d %H:%M') if t.created_at else '',
            ])
        _apply_body_style(ws, data_rows)
        _auto_column_width(ws, headers, data_rows)
        return _wb_to_response(wb, '任务清单')

    @staticmethod
    def export_contributions(project_id, user=None):
        """导出成员贡献记录 Excel"""
        from apps.contributions.models import Contribution
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '贡献记录'
        headers = ['项目', '贡献人', '贡献类型', '贡献内容', '权重', '审核状态',
                   '填写人', '审核人', '审核意见', '统计周期', '创建时间']
        _apply_header_style(ws, headers)

        contributions = Contribution.objects.select_related(
            'project', 'user', 'filled_by', 'reviewer'
        ).filter(project_id=project_id).order_by('-created_at')
        if user is not None:
            from common.project_access import scope_project_queryset

            contributions = scope_project_queryset(
                contributions,
                user,
                project_lookup='project',
            )
        data_rows = []
        for c in contributions:
            data_rows.append([
                c.project.name if c.project else '',
                c.user.name if c.user else '',
                c.get_contribution_type_display(),
                c.content or c.description,
                float(c.weight),
                c.get_status_display(),
                c.filled_by.name if c.filled_by else '',
                c.reviewer.name if c.reviewer else '',
                c.review_opinion,
                c.period,
                c.created_at.strftime('%Y-%m-%d %H:%M') if c.created_at else '',
            ])
        _apply_body_style(ws, data_rows)
        _auto_column_width(ws, headers, data_rows)
        return _wb_to_response(wb, '贡献记录')

    @staticmethod
    def export_ip_applications(user=None):
        """导出知识产权申请总表 Excel"""
        from apps.intellectual_property.models import IntellectualPropertyApplication
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '知识产权申请'
        headers = ['成果名称', '内部编号', '成果类型', '关联项目', '当前状态',
                   '主导撰写人', '申请执行人', '退回次数', '提交日期', '受理日期',
                   '授权日期', '创建时间']
        _apply_header_style(ws, headers)

        applications = IntellectualPropertyApplication.objects.select_related(
            'related_project', 'main_writer', 'applicant_executor'
        ).all().order_by('-created_at')
        if user is not None:
            from apps.intellectual_property.permissions import (
                accessible_ip_applications,
            )

            applications = applications.filter(
                pk__in=accessible_ip_applications(user).values('pk'),
            )
        data_rows = []
        for a in applications:
            data_rows.append([
                a.title,
                a.application_code,
                a.get_ip_type_display(),
                a.related_project.name if a.related_project else '',
                a.get_status_display(),
                a.main_writer.name if a.main_writer else '',
                a.applicant_executor.name if a.applicant_executor else '',
                a.return_count,
                a.submit_date.strftime('%Y-%m-%d') if a.submit_date else '',
                a.accepted_date.strftime('%Y-%m-%d') if a.accepted_date else '',
                a.authorized_date.strftime('%Y-%m-%d') if a.authorized_date else '',
                a.created_at.strftime('%Y-%m-%d %H:%M') if a.created_at else '',
            ])
        _apply_body_style(ws, data_rows)
        _auto_column_width(ws, headers, data_rows)
        return _wb_to_response(wb, '知识产权申请总表')

    @staticmethod
    def export_members(user=None):
        """导出成员列表 Excel"""
        from apps.users.models import User
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '成员列表'
        headers = ['姓名', '邮箱', '手机', '全局角色', '是否学生', '年级', '专业', '状态', '注册时间']
        _apply_header_style(ws, headers)

        users = User.objects.select_related().all().order_by('-date_joined')
        if user is not None:
            from common.project_access import scope_organization_users

            users = scope_organization_users(users, user)
        data_rows = []
        for u in users:
            data_rows.append([
                u.name,
                u.email,
                u.phone,
                u.get_global_role_display(),
                '是' if u.is_student else '否',
                u.grade,
                u.major,
                '启用' if u.is_active else '停用',
                u.date_joined.strftime('%Y-%m-%d %H:%M') if u.date_joined else '',
            ])
        _apply_body_style(ws, data_rows)
        _auto_column_width(ws, headers, data_rows)
        return _wb_to_response(wb, '成员列表')

    @staticmethod
    def export_competitions(
        search='',
        level='',
        status='',
        project_id=None,
        user=None,
    ):
        """按比赛列表当前筛选导出全流程 Excel。"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '比赛列表'
        _apply_header_style(ws, _COMPETITION_EXPORT_HEADERS)

        data_rows = _competition_export_rows(
            search,
            level,
            status,
            project_id,
            user,
        )
        _apply_body_style(ws, data_rows)
        _auto_column_width(ws, _COMPETITION_EXPORT_HEADERS, data_rows)
        return _wb_to_response(wb, '比赛列表')


class CsvExportService:
    """
    CSV 导出服务
    与 ExcelExportService 方法一一对应，输出 CSV 文件流
    使用标准库 csv + io.StringIO 生成，并以 UTF-8 BOM 前缀保证 Excel 正确识别中文编码
    """

    @staticmethod
    def _csv_to_response(headers, data_rows, filename):
        """
        将表头与数据行写入 CSV 并包装为 HttpResponse
        :param headers: 表头列表
        :param data_rows: 数据行列表（每行为列表）
        :param filename: 下载文件名（不含扩展名）
        """
        output = io.StringIO()
        output.write('\ufeff')  # UTF-8 BOM，Excel 兼容
        writer = csv.writer(output)
        writer.writerow(headers)
        writer.writerows(data_rows)
        response = HttpResponse(output.getvalue(), content_type='text/csv')
        # 文件名兼容中文
        response['Content-Disposition'] = f"attachment; filename*=UTF-8''{filename}.csv"
        return response

    @staticmethod
    def export_projects(user=None):
        """导出项目列表 CSV"""
        from apps.projects.models import Project
        headers = ['项目名称', '项目编号', '负责人', '当前阶段', '状态', '开始时间', '预计结束', '创建时间']
        projects = Project.objects.select_related('leader').all().order_by('-created_at')
        if user is not None:
            from common.project_access import scope_project_queryset

            projects = scope_project_queryset(projects, user, project_lookup='')
        data_rows = []
        for p in projects:
            data_rows.append([
                p.name,
                p.code,
                p.leader.name if p.leader else '',
                p.get_current_stage_display(),
                p.get_status_display(),
                p.start_date.strftime('%Y-%m-%d') if p.start_date else '',
                p.planned_end_date.strftime('%Y-%m-%d') if p.planned_end_date else '',
                p.created_at.strftime('%Y-%m-%d %H:%M') if p.created_at else '',
            ])
        return CsvExportService._csv_to_response(headers, data_rows, '项目列表')

    @staticmethod
    def export_finance_budget(user=None):
        """导出经费总表 CSV"""
        from apps.finance.models import FinanceBudget
        headers = ['项目名称', '项目编号', '奖金总额', '其他收入', '已用金额',
                   '待报销', '剩余金额', '经费状态', '统计周期', '更新时间']
        budgets = FinanceBudget.objects.select_related('project').all().order_by('-updated_at')
        if user is not None:
            from common.project_access import scope_project_queryset

            budgets = scope_project_queryset(
                budgets,
                user,
                project_lookup='project',
            )
        data_rows = []
        for b in budgets:
            data_rows.append([
                b.project.name if b.project else '',
                b.project.code if b.project else '',
                float(b.bonus_amount),
                float(b.other_income),
                float(b.used_amount),
                float(b.pending_reimbursement),
                float(b.remaining_amount),
                b.get_status_display(),
                b.period,
                b.updated_at.strftime('%Y-%m-%d %H:%M') if b.updated_at else '',
            ])
        return CsvExportService._csv_to_response(headers, data_rows, '经费总表')

    @staticmethod
    def export_finance_detail(project_id, user=None):
        """导出单项目经费明细 CSV"""
        from apps.finance.models import FinanceExpense
        headers = ['支出标题', '项目名称', '金额', '经办人', '支出日期',
                   '经费类别', '用途说明', '审核人', '创建时间']
        expenses = FinanceExpense.objects.select_related(
            'project', 'spender', 'reviewer'
        ).filter(project_id=project_id).order_by('-expense_date')
        if user is not None:
            from common.project_access import scope_project_queryset

            expenses = scope_project_queryset(
                expenses,
                user,
                project_lookup='project',
            )
        data_rows = []
        for e in expenses:
            data_rows.append([
                e.title,
                e.project.name if e.project else '',
                float(e.amount),
                e.spender.name if e.spender else '',
                e.expense_date.strftime('%Y-%m-%d') if e.expense_date else '',
                e.get_category_display(),
                e.purpose,
                e.reviewer.name if e.reviewer else '',
                e.created_at.strftime('%Y-%m-%d %H:%M') if e.created_at else '',
            ])
        return CsvExportService._csv_to_response(headers, data_rows, '经费明细')

    @staticmethod
    def export_tasks(project_id=None, filters=None, user=None):
        """按任务列表当前筛选导出任务清单 CSV。"""
        headers = [
            '任务标题', '所属项目', '指派给', '协作者', '审核人', '创建者',
            '状态', '优先级', '截止时间', '完成时间', '是否逾期',
            '延期原因', '完成说明', '创建时间',
        ]
        tasks = _task_export_queryset(project_id, filters, user)
        data_rows = []
        for t in tasks:
            data_rows.append([
                t.title,
                t.project.name if t.project else '',
                t.assignee.name if t.assignee else '',
                '、'.join(user.name for user in t.collaborators.all()),
                t.reviewer.name if t.reviewer else '',
                t.creator.name if t.creator else '',
                t.get_status_display(),
                t.get_priority_display(),
                t.deadline.strftime('%Y-%m-%d %H:%M') if t.deadline else '',
                t.completed_at.strftime('%Y-%m-%d %H:%M') if t.completed_at else '',
                '是' if t.is_overdue else '否',
                t.delay_reason,
                t.completion_note,
                t.created_at.strftime('%Y-%m-%d %H:%M') if t.created_at else '',
            ])
        return CsvExportService._csv_to_response(headers, data_rows, '任务清单')

    @staticmethod
    def export_contributions(project_id, user=None):
        """导出成员贡献记录 CSV"""
        from apps.contributions.models import Contribution
        headers = ['项目', '贡献人', '贡献类型', '贡献内容', '权重', '审核状态',
                   '填写人', '审核人', '审核意见', '统计周期', '创建时间']
        contributions = Contribution.objects.select_related(
            'project', 'user', 'filled_by', 'reviewer'
        ).filter(project_id=project_id).order_by('-created_at')
        if user is not None:
            from common.project_access import scope_project_queryset

            contributions = scope_project_queryset(
                contributions,
                user,
                project_lookup='project',
            )
        data_rows = []
        for c in contributions:
            data_rows.append([
                c.project.name if c.project else '',
                c.user.name if c.user else '',
                c.get_contribution_type_display(),
                c.content or c.description,
                float(c.weight),
                c.get_status_display(),
                c.filled_by.name if c.filled_by else '',
                c.reviewer.name if c.reviewer else '',
                c.review_opinion,
                c.period,
                c.created_at.strftime('%Y-%m-%d %H:%M') if c.created_at else '',
            ])
        return CsvExportService._csv_to_response(headers, data_rows, '贡献记录')

    @staticmethod
    def export_ip_applications(user=None):
        """导出知识产权申请总表 CSV"""
        from apps.intellectual_property.models import IntellectualPropertyApplication
        headers = ['成果名称', '内部编号', '成果类型', '关联项目', '当前状态',
                   '主导撰写人', '申请执行人', '退回次数', '提交日期', '受理日期',
                   '授权日期', '创建时间']
        applications = IntellectualPropertyApplication.objects.select_related(
            'related_project', 'main_writer', 'applicant_executor'
        ).all().order_by('-created_at')
        if user is not None:
            from apps.intellectual_property.permissions import (
                accessible_ip_applications,
            )

            applications = applications.filter(
                pk__in=accessible_ip_applications(user).values('pk'),
            )
        data_rows = []
        for a in applications:
            data_rows.append([
                a.title,
                a.application_code,
                a.get_ip_type_display(),
                a.related_project.name if a.related_project else '',
                a.get_status_display(),
                a.main_writer.name if a.main_writer else '',
                a.applicant_executor.name if a.applicant_executor else '',
                a.return_count,
                a.submit_date.strftime('%Y-%m-%d') if a.submit_date else '',
                a.accepted_date.strftime('%Y-%m-%d') if a.accepted_date else '',
                a.authorized_date.strftime('%Y-%m-%d') if a.authorized_date else '',
                a.created_at.strftime('%Y-%m-%d %H:%M') if a.created_at else '',
            ])
        return CsvExportService._csv_to_response(headers, data_rows, '知识产权申请总表')

    @staticmethod
    def export_members(user=None):
        """导出成员列表 CSV"""
        from apps.users.models import User
        headers = ['姓名', '邮箱', '手机', '全局角色', '是否学生', '年级', '专业', '状态', '注册时间']
        users = User.objects.select_related().all().order_by('-date_joined')
        if user is not None:
            from common.project_access import scope_organization_users

            users = scope_organization_users(users, user)
        data_rows = []
        for u in users:
            data_rows.append([
                u.name,
                u.email,
                u.phone,
                u.get_global_role_display(),
                '是' if u.is_student else '否',
                u.grade,
                u.major,
                '启用' if u.is_active else '停用',
                u.date_joined.strftime('%Y-%m-%d %H:%M') if u.date_joined else '',
            ])
        return CsvExportService._csv_to_response(headers, data_rows, '成员列表')

    @staticmethod
    def export_competitions(
        search='',
        level='',
        status='',
        project_id=None,
        user=None,
    ):
        """按比赛列表当前筛选导出全流程 CSV。"""
        data_rows = _competition_export_rows(
            search,
            level,
            status,
            project_id,
            user,
        )
        return CsvExportService._csv_to_response(
            _COMPETITION_EXPORT_HEADERS,
            data_rows,
            '比赛列表',
        )
