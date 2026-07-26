"""
审计日志视图
- OperationLogViewSet: 操作日志只读查询（list/retrieve）+ 模块统计 + 最近日志 + 导出 Excel
权限：老师或管理员（IsTeacherOrAdmin）
"""
import io

from django.db.models import Count, Q
from django.http import HttpResponse
from django.utils import timezone
from datetime import timedelta
from rest_framework.decorators import action
from rest_framework.viewsets import ReadOnlyModelViewSet

from common.response import success_response
from common.mixins import MultiSerializerMixin, MultiPermissionMixin
from common.permissions import IsTeacherOrAdmin
from .models import OperationLog
from .serializers import OperationLogSerializer, OperationLogListSerializer


class OperationLogViewSet(MultiSerializerMixin, MultiPermissionMixin, ReadOnlyModelViewSet):
    """
    操作日志 ViewSet（只读）
    - list: 操作日志列表，支持按 module(object_type)、operator、operation_type、时间范围筛选
    - retrieve: 查看操作日志详情
    - module_stats: 按模块统计操作数
    - recent: 最近 N 条操作日志
    - export: 导出操作日志为 Excel（应用相同筛选条件）
    权限：老师或管理员
    """
    queryset = OperationLog.objects.select_related('operator').all()

    serializer_classes_by_action = {
        'list': OperationLogListSerializer,
        'retrieve': OperationLogSerializer,
    }

    permission_classes_by_action = {
        'list': [IsTeacherOrAdmin],
        'retrieve': [IsTeacherOrAdmin],
        'module_stats': [IsTeacherOrAdmin],
        'recent': [IsTeacherOrAdmin],
        'export': [IsTeacherOrAdmin],
    }

    # 默认权限
    permission_classes = [IsTeacherOrAdmin]

    filterset_fields = ['module', 'operator', 'operation_type', 'is_success']
    search_fields = ['description', 'request_path', 'object_type']
    ordering_fields = ['created_at', 'response_status']

    def get_queryset(self):
        """支持按时间范围、模块名、操作类型、操作人筛选"""
        queryset = super().get_queryset()

        # 按 module（object_type）筛选
        module = self.request.query_params.get('module')
        if module:
            queryset = queryset.filter(module=module)

        # 按 operator（操作人 ID）筛选
        operator = self.request.query_params.get('operator')
        if operator:
            queryset = queryset.filter(operator_id=operator)

        # 按 operation_type（操作类型）筛选
        operation_type = self.request.query_params.get('operation_type')
        if operation_type:
            queryset = queryset.filter(operation_type=operation_type)

        # 按时间范围筛选
        start_date = self.request.query_params.get('start_date')
        end_date = self.request.query_params.get('end_date')
        if start_date:
            queryset = queryset.filter(created_at__date__gte=start_date)
        if end_date:
            queryset = queryset.filter(created_at__date__lte=end_date)

        return queryset

    def list(self, request, *args, **kwargs):
        """操作日志列表"""
        queryset = self.filter_queryset(self.get_queryset())
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = self.get_serializer(queryset, many=True)
        return success_response(serializer.data)

    def retrieve(self, request, *args, **kwargs):
        """操作日志详情"""
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        return success_response(serializer.data)

    @action(detail=False, methods=['get'])
    def module_stats(self, request):
        """
        按模块统计操作数
        GET /api/v1/audit/logs/module_stats/
        可选参数: days（统计最近 N 天的数据，默认 30）
        """
        days = request.query_params.get('days', 30)
        try:
            days = int(days)
        except (ValueError, TypeError):
            days = 30

        # 统计最近 N 天的数据
        start_time = timezone.now() - timedelta(days=days)
        queryset = self.get_queryset().filter(created_at__gte=start_time)

        # 按模块分组统计
        stats = queryset.values('module').annotate(
            total=Count('id'),
            success_count=Count('id', filter=Q(is_success=True)),
            fail_count=Count('id', filter=Q(is_success=False)),
        ).order_by('-total')

        # 按操作类型统计
        type_stats = queryset.values('operation_type').annotate(
            total=Count('id')
        ).order_by('-total')

        result = {
            'days': days,
            'module_stats': list(stats),
            'operation_type_stats': list(type_stats),
            'total': queryset.count(),
        }
        return success_response(result)

    @action(detail=False, methods=['get'])
    def recent(self, request):
        """
        最近 N 条操作日志
        GET /api/v1/audit/logs/recent/
        可选参数: limit（条数，默认 20，最大 100）
        """
        limit = request.query_params.get('limit', 20)
        try:
            limit = int(limit)
        except (ValueError, TypeError):
            limit = 20
        # 限制最大 100 条
        limit = min(max(limit, 1), 100)

        queryset = self.get_queryset()[:limit]
        serializer = OperationLogListSerializer(queryset, many=True)
        return success_response(serializer.data)

    @action(detail=False, methods=['get'])
    def export(self, request):
        """
        导出操作日志为 Excel（应用与 list 相同的筛选条件）
        GET /api/v1/audit/operation-logs/export/
        查询参数: search, module, operator, operation_type, start_date, end_date
        返回: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet
        """
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter

        queryset = self.filter_queryset(self.get_queryset())

        wb = Workbook()
        ws = wb.active
        ws.title = '操作日志'

        headers = [
            'ID', '操作人', '操作类型', '操作模块', '对象类型', '对象ID',
            '请求方法', '请求路径', '响应状态码', '是否成功',
            '操作IP', 'User-Agent', '操作描述', '创建时间',
        ]
        ws.append(headers)

        for log in queryset:
            ws.append([
                log.id,
                log.operator.name if log.operator else '',
                log.get_operation_type_display(),
                log.module,
                log.object_type,
                log.object_id,
                log.request_method,
                log.request_path,
                log.response_status,
                '成功' if log.is_success else '失败',
                log.request_ip or '',
                log.user_agent,
                log.description,
                log.created_at.strftime('%Y-%m-%d %H:%M:%S') if log.created_at else '',
            ])

        # 自适应列宽（粗略）
        for col_idx, header in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = max(len(header) * 2, 14)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        filename = f'operation_logs_{timestamp}.xlsx'
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
